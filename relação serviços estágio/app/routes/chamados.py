from flask import Blueprint, request, jsonify, send_file
import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from app import db
from app.models.chamado import Chamado
from app.schemas.chamado_schema import ChamadoSchema
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

chamados_bp = Blueprint('chamados', __name__, url_prefix='/api')

chamado_schema = ChamadoSchema()
chamados_schema = ChamadoSchema(many=True)

@chamados_bp.route('/', methods=['GET'])
def index():
    return jsonify({"message": "API de Chamados"}), 200

@chamados_bp.route('/chamados', methods=['POST'])
def criar_chamado():
    try:
        data = request.json
        if not data:
            return jsonify({"error": "Dados não fornecidos"}), 400

        campos_obrigatorios = ['data', 'horario', 'estagiario', 'unidade', 'sala', 'professor', 'motivo', 'descricao']
        for campo in campos_obrigatorios:
            if campo not in data:
                return jsonify({"error": f"Campo obrigatório '{campo}' não fornecido"}), 400

        try:
            data_formatada = datetime.strptime(data['data'], '%d/%m/%Y').date()
            horario_formatado = datetime.strptime(data['horario'], '%H:%M:%S').time()
        except ValueError as e:
            return jsonify({"error": f"Erro no formato de data ou horário: {str(e)}"}), 400
        
        novo_chamado = Chamado(
            data=data_formatada,
            horario=horario_formatado,
            estagiario=data['estagiario'],
            unidade=data['unidade'],
            sala=data['sala'],
            professor=data['professor'],
            motivo=data['motivo'],
            descricao=data['descricao']
        )
        
        db.session.add(novo_chamado)
        db.session.commit()
        
        return chamado_schema.dump(novo_chamado), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erro ao criar chamado: {str(e)}"}), 500

@chamados_bp.route('/chamados', methods=['GET'])
def listar_chamados():
    try:
        chamados = Chamado.query.all()
        return chamados_schema.dump(chamados), 200
    except Exception as e:
        return jsonify({"error": f"Erro ao listar chamados: {str(e)}"}), 500

@chamados_bp.route('/chamados/<int:id>', methods=['GET'])
def obter_chamado(id):
    chamado = Chamado.query.get_or_404(id)
    return chamado_schema.dump(chamado), 200

@chamados_bp.route('/chamados/<int:id>', methods=['PUT'])
def atualizar_chamado(id):
    chamado = Chamado.query.get_or_404(id)
    data = request.json
    
    for key, value in data.items():
        if hasattr(chamado, key):
            if key == 'data':
                value = datetime.strptime(value, '%d/%m/%Y').date()
            elif key == 'horario':
                value = datetime.strptime(value, '%H:%M:%S').time()
            setattr(chamado, key, value)
    
    db.session.commit()
    return chamado_schema.dump(chamado), 200

@chamados_bp.route('/chamados/<int:id>', methods=['DELETE'])
def deletar_chamado(id):
    chamado = Chamado.query.get_or_404(id)
    db.session.delete(chamado)
    db.session.commit()
    return '', 204

@chamados_bp.route('/exportar', methods=['GET'])
def exportar_chamados():
    try:
        # Criar novo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chamados"

        # Estilo para cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        
        # Cabeçalhos (removido o Status)
        headers = ['Data', 'Horário', 'Estagiário', 'Unidade', 'Sala', 'Professor', 'Motivo', 'Descrição']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Buscar todos os chamados
        chamados = Chamado.query.all()

        # Preencher dados (removido o status)
        for row, chamado in enumerate(chamados, 2):
            ws.cell(row=row, column=1, value=chamado.data.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=chamado.horario.strftime('%H:%M:%S'))
            ws.cell(row=row, column=3, value=chamado.estagiario)
            ws.cell(row=row, column=4, value=chamado.unidade)
            ws.cell(row=row, column=5, value=chamado.sala)
            ws.cell(row=row, column=6, value=chamado.professor)
            ws.cell(row=row, column=7, value=chamado.motivo)
            ws.cell(row=row, column=8, value=chamado.descricao)

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Salvar arquivo
        data_atual = datetime.now().strftime('%d-%m-%Y')
        filename = f'chamados_{data_atual}.xlsx'
        filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({"error": f"Erro ao exportar chamados: {str(e)}"}), 500